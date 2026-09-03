#!/usr/bin/env python3
"""Capture, convert, replay, and visualize VL53L8CX USART6 frames.

Examples:
  python tof_usart6_receiver.py --port COM3 --out output/tof_capture.xlsx
  python tof_usart6_receiver.py --input output/tof_capture.xlsx --visualize
  python tof_usart6_receiver.py --input output/tof_capture.csv --visualize --loop --speed 2
  python tof_usart6_receiver.py --input output/tof_capture.xlsx --points-out output/tof_points.csv
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime
import json
import math
from pathlib import Path
import sys
import time
from typing import Iterator, List, Protocol


VALID_RESOLUTIONS = (16, 64)
VALID_TARGET_STATUSES = (5, 9)
DEFAULT_DIAGONAL_FOV_DEG = 65.0


@dataclass
class TofFrame:
    frame_id: int
    mcu_ts_ms: int
    resolution: int
    fps: int
    distance_mm: List[int]
    target_status: List[int]
    signal_per_spad: List[int]
    ambient_per_spad: List[int]
    valid: List[int]
    host_ts_ns: int


@dataclass
class FrameRecord:
    sample_idx: int
    unix_ts: float
    raw_line: str
    frame: TofFrame


@dataclass
class PointSample:
    zone_index: int
    x_m: float
    y_m: float
    z_m: float
    distance_mm: int
    target_status: int
    signal_per_spad: int
    ambient_per_spad: int
    color_rgb: tuple[float, float, float]


@dataclass
class InputStats:
    rows_seen: int = 0
    valid_frames: int = 0
    invalid_rows: int = 0


class FrameWriter(Protocol):
    path: Path

    def write(self, record: FrameRecord) -> None: ...

    def close(self) -> None: ...


class PointCloudViewer(Protocol):
    def update(self, points: list[PointSample]) -> bool: ...

    def close(self) -> None: ...


def _parse_int_list(tokens: List[str], start: int, count: int) -> List[int]:
    end = start + count
    if end > len(tokens):
        raise ValueError("not enough numeric tokens")
    values = tokens[start:end]
    if any(value == "" for value in values):
        raise ValueError("empty numeric token")
    return [int(value) for value in values]


def parse_tof_line(line: str, host_ts_ns: int | None = None) -> TofFrame:
    tokens = [token.strip() for token in line.strip().split(",")]
    if len(tokens) < 10:
        raise ValueError("line too short")
    if tokens[0] != "TOF":
        raise ValueError("not a TOF frame")

    frame_id = int(tokens[1])
    mcu_ts_ms = int(tokens[2])
    resolution = int(tokens[3])
    fps = int(tokens[4])
    if resolution not in VALID_RESOLUTIONS:
        raise ValueError(f"unsupported resolution: {resolution}")
    if fps <= 0:
        raise ValueError(f"invalid fps: {fps}")

    idx = 5
    fields: dict[str, List[int]] = {}
    for marker in ("D", "S", "G", "A"):
        if idx >= len(tokens) or tokens[idx] != marker:
            raise ValueError(f"missing {marker} marker")
        idx += 1
        fields[marker] = _parse_int_list(tokens, idx, resolution)
        idx += resolution

    if idx < len(tokens):
        if tokens[idx] != "V":
            raise ValueError("missing V marker")
        idx += 1
        valid = _parse_int_list(tokens, idx, resolution)
        idx += resolution
    else:
        valid = [1 if status in VALID_TARGET_STATUSES else 0 for status in fields["S"]]

    if idx != len(tokens):
        raise ValueError(f"unexpected trailing fields: {len(tokens) - idx}")
    if any(value not in (0, 1) for value in valid):
        raise ValueError("valid array must contain only 0 or 1")

    return TofFrame(
        frame_id=frame_id,
        mcu_ts_ms=mcu_ts_ms,
        resolution=resolution,
        fps=fps,
        distance_mm=fields["D"],
        target_status=fields["S"],
        signal_per_spad=fields["G"],
        ambient_per_spad=fields["A"],
        valid=valid,
        host_ts_ns=time.monotonic_ns() if host_ts_ns is None else host_ts_ns,
    )


def frame_to_tof_line(frame: TofFrame) -> str:
    parts: list[str] = [
        "TOF",
        str(frame.frame_id),
        str(frame.mcu_ts_ms),
        str(frame.resolution),
        str(frame.fps),
    ]
    for marker, values in (
        ("D", frame.distance_mm),
        ("S", frame.target_status),
        ("G", frame.signal_per_spad),
        ("A", frame.ambient_per_spad),
        ("V", frame.valid),
    ):
        parts.append(marker)
        parts.extend(str(value) for value in values)
    return ",".join(parts)


def frame_to_features(frame: TofFrame) -> dict:
    valid_mask = [value == 1 for value in frame.valid]
    valid_distances = [
        distance
        for distance, status, is_valid in zip(
            frame.distance_mm, frame.target_status, valid_mask
        )
        if is_valid and status in VALID_TARGET_STATUSES
    ]
    valid_count = len(valid_distances)
    total_count = len(valid_mask)
    return {
        "valid_count": valid_count,
        "total_count": total_count,
        "valid_ratio": valid_count / total_count if total_count else 0.0,
        "min_distance_mm": min(valid_distances) if valid_distances else -1,
        "avg_distance_mm": (
            sum(valid_distances) / valid_count if valid_count else -1.0
        ),
        "avg_signal": sum(frame.signal_per_spad) / len(frame.signal_per_spad),
        "avg_ambient": sum(frame.ambient_per_spad) / len(frame.ambient_per_spad),
    }


def _format_grid(title: str, values: List[int], resolution: int) -> list[str]:
    side = math.isqrt(resolution)
    if side * side != resolution:
        return [f"{title}: " + " ".join(str(value) for value in values)]

    lines = [f"{title}:"]
    for row in range(side):
        start = row * side
        lines.append("  " + " ".join(f"{value:>5}" for value in values[start:start + side]))
    return lines


def _print_frame_grid(frame: TofFrame) -> None:
    for values, title in (
        (frame.distance_mm, "distance_mm"),
        (frame.target_status, "status"),
        (frame.valid, "valid"),
    ):
        for line in _format_grid(title, values, frame.resolution):
            print(line)


class JsonlFrameWriter:
    def __init__(self, path: Path) -> None:
        self.path = path
        self._file = path.open("w", encoding="utf-8")

    def write(self, record: FrameRecord) -> None:
        data = {
            "frame": record.frame.__dict__,
            "features": frame_to_features(record.frame),
        }
        self._file.write(json.dumps(data, ensure_ascii=True) + "\n")

    def close(self) -> None:
        self._file.close()


class CsvFrameWriter:
    HEADERS = (
        "sample_idx",
        "host_ts_ns",
        "t_unix",
        "time",
        "frame_id",
        "mcu_ts_ms",
        "resolution",
        "fps",
        "tof_line",
    )

    def __init__(self, path: Path) -> None:
        self.path = path
        self._file = path.open("w", encoding="utf-8-sig", newline="")
        self._writer = csv.writer(self._file)
        self._writer.writerow(self.HEADERS)

    def write(self, record: FrameRecord) -> None:
        frame = record.frame
        self._writer.writerow(
            (
                record.sample_idx,
                frame.host_ts_ns,
                f"{record.unix_ts:.9f}",
                datetime.fromtimestamp(record.unix_ts).isoformat(timespec="milliseconds"),
                frame.frame_id,
                frame.mcu_ts_ms,
                frame.resolution,
                frame.fps,
                record.raw_line,
            )
        )

    def close(self) -> None:
        self._file.close()


class XlsxFrameWriter:
    HEADERS = CsvFrameWriter.HEADERS

    def __init__(self, path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.cell import WriteOnlyCell
            from openpyxl.styles import Alignment, Font, PatternFill
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "XLSX support requires openpyxl: python -m pip install openpyxl"
            ) from exc

        self.path = path
        self._WriteOnlyCell = WriteOnlyCell
        self._workbook = Workbook(write_only=True)
        self._sheet = self._workbook.create_sheet("in")
        self._sheet.freeze_panes = "A2"
        self._sheet.column_dimensions["A"].width = 12
        self._sheet.column_dimensions["B"].width = 22
        self._sheet.column_dimensions["C"].width = 18
        self._sheet.column_dimensions["D"].width = 24
        self._sheet.column_dimensions["E"].width = 12
        self._sheet.column_dimensions["F"].width = 14
        self._sheet.column_dimensions["G"].width = 12
        self._sheet.column_dimensions["H"].width = 8
        self._sheet.column_dimensions["I"].width = 100

        header = []
        for value in self.HEADERS:
            cell = WriteOnlyCell(self._sheet, value=value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
            header.append(cell)
        self._sheet.append(header)

    def write(self, record: FrameRecord) -> None:
        frame = record.frame
        host_timestamp = self._WriteOnlyCell(
            self._sheet, value=str(frame.host_ts_ns)
        )
        host_timestamp.number_format = "@"
        timestamp = self._WriteOnlyCell(
            self._sheet, value=datetime.fromtimestamp(record.unix_ts)
        )
        timestamp.number_format = "yyyy-mm-dd hh:mm:ss.000"
        self._sheet.append(
            (
                record.sample_idx,
                host_timestamp,
                record.unix_ts,
                timestamp,
                frame.frame_id,
                frame.mcu_ts_ms,
                frame.resolution,
                frame.fps,
                record.raw_line,
            )
        )

    def close(self) -> None:
        self._workbook.save(self.path)


class PointCsvWriter:
    def __init__(self, path: Path) -> None:
        self.path = path
        path.parent.mkdir(parents=True, exist_ok=True)
        self._file = path.open("w", encoding="utf-8-sig", newline="")
        self._writer = csv.writer(self._file)
        self._writer.writerow(
            (
                "frame_id",
                "mcu_ts_ms",
                "host_ts_ns",
                "zone_index",
                "x_m",
                "y_m",
                "z_m",
                "distance_mm",
                "target_status",
                "signal_per_spad",
                "ambient_per_spad",
            )
        )

    def write(self, frame: TofFrame, points: list[PointSample]) -> None:
        for point in points:
            self._writer.writerow(
                (
                    frame.frame_id,
                    frame.mcu_ts_ms,
                    frame.host_ts_ns,
                    point.zone_index,
                    f"{point.x_m:.6f}",
                    f"{point.y_m:.6f}",
                    f"{point.z_m:.6f}",
                    point.distance_mm,
                    point.target_status,
                    point.signal_per_spad,
                    point.ambient_per_spad,
                )
            )

    def close(self) -> None:
        self._file.close()


def create_frame_writer(path_value: str | None) -> FrameWriter | None:
    if not path_value:
        return None
    path = Path(path_value)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix in (".jsonl", ".json"):
        return JsonlFrameWriter(path)
    if suffix == ".csv":
        return CsvFrameWriter(path)
    if suffix == ".xlsx":
        return XlsxFrameWriter(path)
    raise ValueError("output extension must be .jsonl, .csv, or .xlsx")


def _host_timestamp_from_row(row: dict[str, object]) -> int:
    host_value = row.get("host_ts_ns")
    if host_value not in (None, ""):
        return int(host_value)
    unix_value = row.get("t_unix")
    if unix_value not in (None, ""):
        return int(float(unix_value) * 1_000_000_000)
    return time.monotonic_ns()


def _unix_timestamp_from_row(row: dict[str, object]) -> float:
    value = row.get("t_unix")
    if value not in (None, ""):
        return float(value)
    return time.time()


def _record_from_row(
    row: dict[str, object], fallback_sample_idx: int
) -> FrameRecord:
    raw_value = row.get("tof_line") or row.get("raw_line") or row.get("line")
    if not isinstance(raw_value, str) or not raw_value.strip():
        raise ValueError("missing tof_line column value")
    raw_line = raw_value.strip()
    host_ts_ns = _host_timestamp_from_row(row)
    frame = parse_tof_line(raw_line, host_ts_ns=host_ts_ns)
    sample_value = row.get("sample_idx")
    sample_idx = int(sample_value) if sample_value not in (None, "") else fallback_sample_idx
    return FrameRecord(
        sample_idx=sample_idx,
        unix_ts=_unix_timestamp_from_row(row),
        raw_line=raw_line,
        frame=frame,
    )


def _report_bad_row(path: Path, row_number: int, exc: Exception, stats: InputStats) -> None:
    stats.invalid_rows += 1
    if stats.invalid_rows <= 5:
        print(f"skip invalid row {path.name}:{row_number}: {exc}", file=sys.stderr)
    elif stats.invalid_rows == 6:
        print("additional invalid-row messages suppressed", file=sys.stderr)


def iter_csv_records(path: Path, stats: InputStats) -> Iterator[FrameRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        first_line = file.readline()
        file.seek(0)
        if first_line.lstrip().startswith("TOF,"):
            for row_number, raw_line in enumerate(file, start=1):
                stats.rows_seen += 1
                try:
                    line = raw_line.strip()
                    frame = parse_tof_line(line)
                    stats.valid_frames += 1
                    yield FrameRecord(row_number - 1, time.time(), line, frame)
                except Exception as exc:
                    _report_bad_row(path, row_number, exc, stats)
            return

        reader = csv.DictReader(file)
        if not reader.fieldnames:
            raise ValueError("CSV file has no header")
        reader.fieldnames = [str(name).strip().lower() for name in reader.fieldnames]
        for row_number, row in enumerate(reader, start=2):
            stats.rows_seen += 1
            normalized = {str(key).strip().lower(): value for key, value in row.items()}
            try:
                record = _record_from_row(normalized, row_number - 2)
                stats.valid_frames += 1
                yield record
            except Exception as exc:
                _report_bad_row(path, row_number, exc, stats)


def iter_xlsx_records(
    path: Path, sheet_name: str | None, stats: InputStats
) -> Iterator[FrameRecord]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX support requires openpyxl: python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"sheet {sheet_name!r} not found; available: {workbook.sheetnames}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            return
        headers = [
            str(value).strip().lower() if value is not None else ""
            for value in header_values
        ]
        if "tof_line" not in headers and "raw_line" not in headers and "line" not in headers:
            raise ValueError("XLSX sheet must contain a tof_line, raw_line, or line column")

        for row_number, values in enumerate(rows, start=2):
            stats.rows_seen += 1
            row = {headers[index]: value for index, value in enumerate(values) if headers[index]}
            try:
                record = _record_from_row(row, row_number - 2)
                stats.valid_frames += 1
                yield record
            except Exception as exc:
                _report_bad_row(path, row_number, exc, stats)
    finally:
        workbook.close()


def iter_jsonl_records(path: Path, stats: InputStats) -> Iterator[FrameRecord]:
    with path.open("r", encoding="utf-8") as file:
        for row_number, line in enumerate(file, start=1):
            stats.rows_seen += 1
            try:
                data = json.loads(line)
                frame_data = data.get("frame", data)
                unchecked_frame = TofFrame(**frame_data)
                raw_line = frame_to_tof_line(unchecked_frame)
                frame = parse_tof_line(
                    raw_line, host_ts_ns=unchecked_frame.host_ts_ns
                )
                stats.valid_frames += 1
                yield FrameRecord(row_number - 1, time.time(), raw_line, frame)
            except Exception as exc:
                _report_bad_row(path, row_number, exc, stats)


def iter_input_records(
    path: Path, sheet_name: str | None, stats: InputStats
) -> Iterator[FrameRecord]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_records(path, stats)
    elif suffix == ".xlsx":
        yield from iter_xlsx_records(path, sheet_name, stats)
    elif suffix in (".jsonl", ".json"):
        yield from iter_jsonl_records(path, stats)
    else:
        raise ValueError("input extension must be .csv, .xlsx, .jsonl, or .json")


def _oriented_coordinates(
    row: int,
    col: int,
    side: int,
    rotation_deg: int,
    flip_x: bool,
    flip_y: bool,
) -> tuple[int, int]:
    if flip_x:
        col = side - 1 - col
    if flip_y:
        row = side - 1 - row
    if rotation_deg == 90:
        return col, side - 1 - row
    if rotation_deg == 180:
        return side - 1 - row, side - 1 - col
    if rotation_deg == 270:
        return side - 1 - col, row
    return row, col


def build_zone_rays(
    resolution: int,
    diagonal_fov_deg: float,
    rotation_deg: int,
    flip_x: bool,
    flip_y: bool,
) -> list[tuple[float, float, float]]:
    side = math.isqrt(resolution)
    if side * side != resolution:
        raise ValueError(f"resolution is not square: {resolution}")
    if not 0.0 < diagonal_fov_deg < 180.0:
        raise ValueError("diagonal FOV must be between 0 and 180 degrees")

    diagonal_tangent = math.tan(math.radians(diagonal_fov_deg) / 2.0)
    axis_tangent = diagonal_tangent / math.sqrt(2.0)
    focal_length = side / (2.0 * axis_tangent)
    center = side / 2.0
    rays: list[tuple[float, float, float]] = []
    for raw_row in range(side):
        for raw_col in range(side):
            row, col = _oriented_coordinates(
                raw_row, raw_col, side, rotation_deg, flip_x, flip_y
            )
            x = (col + 0.5 - center) / focal_length
            y = (row + 0.5 - center) / focal_length
            norm = math.sqrt(x * x + y * y + 1.0)
            rays.append((x / norm, y / norm, 1.0 / norm))
    return rays


def _distance_color(distance_mm: int, min_mm: int, max_mm: int) -> tuple[float, float, float]:
    span = max(max_mm - min_mm, 1)
    value = min(max((distance_mm - min_mm) / span, 0.0), 1.0)
    if value < 0.5:
        t = value * 2.0
        return (0.1, t, 1.0 - 0.4 * t)
    t = (value - 0.5) * 2.0
    return (t, 1.0 - 0.7 * t, 0.6 * (1.0 - t))


def frame_to_point_cloud(
    frame: TofFrame,
    rays: list[tuple[float, float, float]],
    min_distance_mm: int,
    max_distance_mm: int,
) -> list[PointSample]:
    if len(rays) != frame.resolution:
        raise ValueError("ray count does not match frame resolution")
    points: list[PointSample] = []
    for index, (distance, status, valid) in enumerate(
        zip(frame.distance_mm, frame.target_status, frame.valid)
    ):
        if valid != 1 or status not in VALID_TARGET_STATUSES:
            continue
        if distance < min_distance_mm or distance > max_distance_mm:
            continue
        distance_m = distance / 1000.0
        ray_x, ray_y, ray_z = rays[index]
        points.append(
            PointSample(
                zone_index=index,
                x_m=distance_m * ray_x,
                y_m=distance_m * ray_y,
                z_m=distance_m * ray_z,
                distance_mm=distance,
                target_status=status,
                signal_per_spad=frame.signal_per_spad[index],
                ambient_per_spad=frame.ambient_per_spad[index],
                color_rgb=_distance_color(distance, min_distance_mm, max_distance_mm),
            )
        )
    return points


class Open3dPointCloudViewer:
    def __init__(self, point_size: float) -> None:
        try:
            import numpy as np
            import open3d as o3d
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "Dynamic point-cloud visualization requires numpy and open3d. "
                "Install them with: python -m pip install numpy open3d"
            ) from exc

        self._np = np
        self._o3d = o3d
        self._visualizer = o3d.visualization.Visualizer()
        if not self._visualizer.create_window(
            window_name="VL53L8CX Dynamic Point Cloud", width=1100, height=760
        ):
            raise RuntimeError("failed to create Open3D window")
        self._cloud = o3d.geometry.PointCloud()
        self._geometry_added = False
        render_options = self._visualizer.get_render_option()
        render_options.point_size = point_size
        render_options.background_color = np.asarray([0.04, 0.05, 0.06])

    def update(self, points: list[PointSample]) -> bool:
        if not points and not self._geometry_added:
            is_open = self._visualizer.poll_events()
            self._visualizer.update_renderer()
            return bool(is_open)
        xyz = self._np.asarray(
            [(point.x_m, point.y_m, point.z_m) for point in points], dtype=float
        ).reshape((-1, 3))
        colors = self._np.asarray(
            [point.color_rgb for point in points], dtype=float
        ).reshape((-1, 3))
        self._cloud.points = self._o3d.utility.Vector3dVector(xyz)
        self._cloud.colors = self._o3d.utility.Vector3dVector(colors)
        if not self._geometry_added:
            self._visualizer.add_geometry(self._cloud)
            axes = self._o3d.geometry.TriangleMesh.create_coordinate_frame(size=0.2)
            self._visualizer.add_geometry(axes, reset_bounding_box=False)
            self._geometry_added = True
            view = self._visualizer.get_view_control()
            view.set_front([0.0, -0.35, -1.0])
            view.set_up([0.0, -1.0, 0.0])
            view.set_lookat([0.0, 0.0, 0.6])
            view.set_zoom(0.55)
        else:
            self._visualizer.update_geometry(self._cloud)
        is_open = self._visualizer.poll_events()
        self._visualizer.update_renderer()
        return bool(is_open)

    def close(self) -> None:
        self._visualizer.destroy_window()


class MatplotlibPointCloudViewer:
    def __init__(self, point_size: float, max_distance_m: float) -> None:
        try:
            import matplotlib.pyplot as plt
            import numpy as np
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "Matplotlib visualization requires numpy and matplotlib. "
                "Install them with: python -m pip install numpy matplotlib"
            ) from exc

        self._plt = plt
        self._np = np
        self._closed = False
        self._interactive = "agg" not in plt.get_backend().lower()
        self._point_size = max(point_size * 2.0, 4.0)
        self._figure = plt.figure("VL53L8CX Dynamic Point Cloud", figsize=(10, 7))
        self._axes = self._figure.add_subplot(111, projection="3d")
        self._scatter = self._axes.scatter([], [], [], s=self._point_size)
        self._title = self._axes.set_title("Waiting for a valid TOF frame")
        lateral_limit = max(max_distance_m * 0.55, 0.2)
        self._axes.set_xlim(-lateral_limit, lateral_limit)
        self._axes.set_ylim(lateral_limit, -lateral_limit)
        self._axes.set_zlim(0.0, max_distance_m)
        self._axes.set_xlabel("X right (m)")
        self._axes.set_ylabel("Y down (m)")
        self._axes.set_zlabel("Z forward (m)")
        self._axes.set_box_aspect((1.0, 1.0, 1.5))
        self._axes.view_init(elev=-20.0, azim=-90.0)
        self._axes.grid(True, alpha=0.3)
        self._figure.canvas.mpl_connect("close_event", self._on_close)
        if self._interactive:
            plt.ion()
            plt.show(block=False)

    def _on_close(self, _event) -> None:
        self._closed = True

    def update(self, points: list[PointSample]) -> bool:
        if self._closed or not self._plt.fignum_exists(self._figure.number):
            return False
        x_values = [point.x_m for point in points]
        y_values = [point.y_m for point in points]
        z_values = [point.z_m for point in points]
        colors = [point.color_rgb for point in points]
        self._scatter._offsets3d = (x_values, y_values, z_values)
        self._scatter.set_sizes(
            self._np.full(len(points), self._point_size, dtype=float)
        )
        if colors:
            self._scatter.set_color(colors)
        self._title.set_text(f"VL53L8CX point cloud: {len(points)} valid points")
        if self._interactive:
            self._figure.canvas.draw_idle()
            self._figure.canvas.flush_events()
            self._plt.pause(0.001)
        else:
            self._figure.canvas.draw()
        return not self._closed

    def close(self) -> None:
        if self._plt.fignum_exists(self._figure.number):
            self._plt.close(self._figure)


def create_point_cloud_viewer(
    backend: str, point_size: float, max_distance_m: float
) -> PointCloudViewer:
    backends = ("matplotlib", "open3d") if backend == "auto" else (backend,)
    errors: list[str] = []
    for candidate in backends:
        try:
            if candidate == "matplotlib":
                return MatplotlibPointCloudViewer(point_size, max_distance_m)
            if candidate == "open3d":
                return Open3dPointCloudViewer(point_size)
        except RuntimeError as exc:
            errors.append(f"{candidate}: {exc}")
    raise RuntimeError("; ".join(errors))


def _load_serial_modules():
    try:
        import serial
        from serial.tools import list_ports
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Serial capture requires pyserial: python -m pip install pyserial"
        ) from exc
    return serial, list_ports


def _choose_serial_port(cli_port: str | None) -> str:
    if cli_port:
        return cli_port
    _, list_ports = _load_serial_modules()
    ports = sorted(list(list_ports.comports()), key=lambda item: item.device)
    if not ports:
        raise RuntimeError("no serial ports found")

    print("Available serial ports:")
    for index, port in enumerate(ports, start=1):
        print(f"  {index}. {port.device} - {port.description or 'unknown'}")
    while True:
        choice = input("Select serial port number: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(ports):
            return ports[int(choice) - 1].device
        print(f"Enter a number from 1 to {len(ports)}.")


def _print_progress(record: FrameRecord, points: list[PointSample], count: int) -> None:
    features = frame_to_features(record.frame)
    print(
        f"frame={record.frame.frame_id} captured={count} "
        f"valid={features['valid_count']}/{features['total_count']} "
        f"points={len(points)} min={features['min_distance_mm']}mm "
        f"avg={features['avg_distance_mm']:.1f}mm"
    )


def _playback_delay_seconds(
    previous_frame: TofFrame | None,
    frame: TofFrame,
    override_fps: float,
    speed: float,
) -> float:
    if override_fps > 0.0:
        return 1.0 / override_fps / speed
    if previous_frame is not None:
        delta_ms = (frame.mcu_ts_ms - previous_frame.mcu_ts_ms) & 0xFFFFFFFF
        if 0 < delta_ms <= 1000:
            return delta_ms / 1000.0 / speed
    return 1.0 / max(frame.fps, 1) / speed


def _make_rays_for_frame(frame: TofFrame, args: argparse.Namespace):
    return build_zone_rays(
        frame.resolution,
        args.diagonal_fov_deg,
        args.rotate,
        args.flip_x,
        args.flip_y,
    )


def run_file_mode(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    if not input_path.is_file():
        raise FileNotFoundError(f"input file not found: {input_path}")
    if args.out and Path(args.out).resolve() == input_path.resolve():
        raise ValueError("--out must not overwrite --input")
    if args.points_out and Path(args.points_out).resolve() == input_path.resolve():
        raise ValueError("--points-out must not overwrite --input")
    viewer = (
        create_point_cloud_viewer(
            args.viewer, args.point_size, args.max_distance_mm / 1000.0
        )
        if args.visualize
        else None
    )
    frame_writer = create_frame_writer(args.out)
    point_writer = PointCsvWriter(Path(args.points_out)) if args.points_out else None
    loop_index = 0
    total_displayed = 0
    keep_running = True
    try:
        while keep_running:
            stats = InputStats()
            previous_frame: TofFrame | None = None
            rays_by_resolution: dict[int, list[tuple[float, float, float]]] = {}
            displayed_this_loop = 0
            for record in iter_input_records(input_path, args.sheet, stats):
                if args.max_frames > 0 and displayed_this_loop >= args.max_frames:
                    break
                frame = record.frame
                if frame.resolution not in rays_by_resolution:
                    rays_by_resolution[frame.resolution] = _make_rays_for_frame(frame, args)
                points = frame_to_point_cloud(
                    frame,
                    rays_by_resolution[frame.resolution],
                    args.min_distance_mm,
                    args.max_distance_mm,
                )

                if viewer is not None:
                    if previous_frame is not None:
                        time.sleep(
                            _playback_delay_seconds(
                                previous_frame, frame, args.playback_fps, args.speed
                            )
                        )
                    if not viewer.update(points):
                        keep_running = False
                        break
                if loop_index == 0:
                    if frame_writer is not None:
                        frame_writer.write(record)
                    if point_writer is not None:
                        point_writer.write(frame, points)

                displayed_this_loop += 1
                total_displayed += 1
                if args.print_every > 0 and displayed_this_loop % args.print_every == 0:
                    _print_progress(record, points, displayed_this_loop)
                    if args.print_grid:
                        _print_frame_grid(frame)
                previous_frame = frame

            print(
                f"file pass={loop_index + 1} rows={stats.rows_seen} "
                f"frames={displayed_this_loop} invalid={stats.invalid_rows}"
            )
            loop_index += 1
            if not args.loop or displayed_this_loop == 0:
                break
    except KeyboardInterrupt:
        print("playback stopped")
    finally:
        if viewer is not None:
            viewer.close()
        if frame_writer is not None:
            frame_writer.close()
            print(f"frames saved: {frame_writer.path}")
        if point_writer is not None:
            point_writer.close()
            print(f"point cloud saved: {point_writer.path}")
    print(f"playback complete: displayed={total_displayed}")


def run_serial_mode(args: argparse.Namespace) -> None:
    serial, _ = _load_serial_modules()
    port = _choose_serial_port(args.port)
    output_value = args.out or "tof_frames.jsonl"
    viewer = (
        create_point_cloud_viewer(
            args.viewer, args.point_size, args.max_distance_mm / 1000.0
        )
        if args.visualize
        else None
    )
    frame_writer = create_frame_writer(output_value)
    point_writer = PointCsvWriter(Path(args.points_out)) if args.points_out else None
    rays_by_resolution: dict[int, list[tuple[float, float, float]]] = {}
    captured = 0
    invalid_lines = 0
    try:
        with serial.Serial(port, args.baud, timeout=1) as serial_port:
            print(f"capturing {port} at {args.baud} baud; press Ctrl+C to stop")
            while args.max_frames <= 0 or captured < args.max_frames:
                raw = serial_port.readline()
                if not raw:
                    continue
                host_ts_ns = time.monotonic_ns()
                unix_ts = time.time()
                try:
                    line = raw.decode("utf-8").strip()
                    frame = parse_tof_line(line, host_ts_ns=host_ts_ns)
                    record = FrameRecord(captured, unix_ts, line, frame)
                    if frame.resolution not in rays_by_resolution:
                        rays_by_resolution[frame.resolution] = _make_rays_for_frame(frame, args)
                    points = frame_to_point_cloud(
                        frame,
                        rays_by_resolution[frame.resolution],
                        args.min_distance_mm,
                        args.max_distance_mm,
                    )
                    if frame_writer is not None:
                        frame_writer.write(record)
                    if point_writer is not None:
                        point_writer.write(frame, points)
                    captured += 1
                    if viewer is not None and not viewer.update(points):
                        break
                    if args.print_every > 0 and captured % args.print_every == 0:
                        _print_progress(record, points, captured)
                        if args.print_grid:
                            _print_frame_grid(frame)
                except Exception as exc:
                    invalid_lines += 1
                    if invalid_lines <= 5:
                        print(f"parse error: {exc}", file=sys.stderr)
                    elif invalid_lines == 6:
                        print("additional parse-error messages suppressed", file=sys.stderr)
    except KeyboardInterrupt:
        print("capture stopped")
    finally:
        if viewer is not None:
            viewer.close()
        if frame_writer is not None:
            frame_writer.close()
            print(f"frames saved: {frame_writer.path}")
        if point_writer is not None:
            point_writer.close()
            print(f"point cloud saved: {point_writer.path}")
    print(f"capture complete: frames={captured} invalid_lines={invalid_lines}")


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Capture or replay VL53L8CX frames and generate dynamic point clouds."
    )
    source = parser.add_mutually_exclusive_group()
    source.add_argument("--input", help="Replay .csv, .xlsx, .jsonl, or .json frames")
    source.add_argument("--port", help="Serial port, e.g. /dev/ttyUSB0 or COM3")
    parser.add_argument("--baud", type=int, default=921600)
    parser.add_argument(
        "--out",
        help="Captured/converted frame output; extension selects .jsonl, .csv, or .xlsx",
    )
    parser.add_argument("--sheet", help="XLSX sheet name; defaults to the first sheet")
    parser.add_argument("--points-out", help="Export valid XYZ points to a long-form CSV")
    parser.add_argument("--visualize", action="store_true", help="Show dynamic 3D point cloud")
    parser.add_argument(
        "--viewer",
        choices=("auto", "matplotlib", "open3d"),
        default="auto",
        help="Point-cloud viewer backend; auto prefers Matplotlib",
    )
    parser.add_argument("--loop", action="store_true", help="Loop file playback until stopped")
    parser.add_argument("--speed", type=float, default=1.0, help="Playback speed multiplier")
    parser.add_argument(
        "--playback-fps",
        type=float,
        default=0.0,
        help="Override playback FPS; 0 uses MCU timestamps/frame FPS",
    )
    parser.add_argument("--max-frames", type=int, default=0, help="Stop after N frames; 0 is unlimited")
    parser.add_argument("--print-every", type=int, default=30)
    parser.add_argument("--print-grid", action="store_true")
    parser.add_argument("--min-distance-mm", type=int, default=120)
    parser.add_argument("--max-distance-mm", type=int, default=1500)
    parser.add_argument("--diagonal-fov-deg", type=float, default=DEFAULT_DIAGONAL_FOV_DEG)
    parser.add_argument("--rotate", type=int, choices=(0, 90, 180, 270), default=0)
    parser.add_argument("--flip-x", action="store_true")
    parser.add_argument("--flip-y", action="store_true")
    parser.add_argument("--point-size", type=float, default=10.0)
    return parser


def _validate_args(args: argparse.Namespace) -> None:
    if args.speed <= 0.0:
        raise ValueError("--speed must be greater than zero")
    if args.playback_fps < 0.0:
        raise ValueError("--playback-fps cannot be negative")
    if args.max_frames < 0:
        raise ValueError("--max-frames cannot be negative")
    if args.min_distance_mm < 0 or args.max_distance_mm <= args.min_distance_mm:
        raise ValueError("distance range is invalid")
    if args.loop and not args.input:
        raise ValueError("--loop is only available with --input")


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()
    try:
        _validate_args(args)
        if args.input:
            run_file_mode(args)
        else:
            run_serial_mode(args)
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc


if __name__ == "__main__":
    main()
